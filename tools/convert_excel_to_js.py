from __future__ import annotations

import json
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = ROOT / "data"
OUTPUT = ROOT / "web" / "survey-data.js"


def clean_header(value: object, index: int) -> str:
    text = str(value).strip() if value is not None else ""
    return text or f"Column {index}"


def serialize(value: object) -> object:
    if isinstance(value, datetime):
        return value.isoformat(sep=" ", timespec="seconds")
    if isinstance(value, date):
        return value.isoformat()
    return value


def workbook_to_payload(path: Path) -> dict:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheets = []

    for worksheet in workbook.worksheets:
        rows = [
            list(row)
            for row in worksheet.iter_rows(values_only=True)
            if any(cell is not None for cell in row)
        ]
        if not rows:
            sheets.append({"name": worksheet.title, "columns": [], "rows": []})
            continue

        columns = [clean_header(value, index) for index, value in enumerate(rows[0], start=1)]
        records = []
        for row in rows[1:]:
            record = {}
            for index, column in enumerate(columns):
                value = row[index] if index < len(row) else None
                record[column] = serialize(value)
            records.append(record)

        sheets.append({"name": worksheet.title, "columns": columns, "rows": records})

    return {
        "fileName": path.name,
        "relativePath": str(path.relative_to(ROOT)),
        "sheets": sheets,
    }


def main() -> None:
    files = sorted(DATA_DIR.glob("*.xlsx"))
    payload = {
        "generatedAt": datetime.now().isoformat(sep=" ", timespec="seconds"),
        "sources": [workbook_to_payload(path) for path in files],
    }
    content = "window.SURVEY_DATA = "
    content += json.dumps(payload, ensure_ascii=False, indent=2)
    content += ";\n"
    OUTPUT.write_text(content, encoding="utf-8")
    print(f"Wrote {OUTPUT.relative_to(ROOT)} with {len(files)} source files")


if __name__ == "__main__":
    main()
